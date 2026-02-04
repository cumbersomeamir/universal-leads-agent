#!/usr/bin/env python3
"""
Run Apify Website Content Crawler for Rightmove London commercial to let.
Requires APIFY_TOKEN, APIFY_API_TOKEN, or Apify_Website_API_KEY in environment (or .env).
Output: rightmove_london_commercial.xlsx in this folder.
"""

import json
import logging
import os
import re
import sys
import time
from pathlib import Path

# Logging: timestamp + level + message, to stderr so terminal shows it
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    stream=sys.stderr,
    force=True,
)
log = logging.getLogger(__name__)

# Load .env from this folder if present
_script_dir = Path(__file__).resolve().parent
_env_file = _script_dir / ".env"
if _env_file.exists():
    try:
        with open(_env_file) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    except Exception:
        pass

try:
    from apify_client import ApifyClient
except ImportError:
    print("Install: pip install apify-client", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RIGHTMOVE_BASE = (
    "https://www.rightmove.co.uk/commercial-property-to-let/find.html"
    "?searchLocation=London"
    "&useLocationIdentifier=true"
    "&locationIdentifier=REGION%5E87490"
    "&radius=40.0"
    "&_includeLetAgreed=on"
)

# 24 results per page; ~17,393 listings ≈ 725 pages – use 750 to cover all
MAX_START_PAGES = 750
SAVE_EVERY_N_PROPERTIES = 10
# Chunk size per Apify run – keep small to stay under Apify memory limit (8192MB total)
URLS_PER_CHUNK = 5
MAX_PAGES_PER_CHUNK = 150
CHUNK_RETRY_WAIT_SEC = 90
CHUNK_MAX_RETRIES = 10

HEADERS = ["Property Name", "Address", "Link to Property", "Price", "Size", "Property Details", "Phone Number"]


def build_start_urls():
    urls = [{"url": RIGHTMOVE_BASE}]
    for i in range(1, MAX_START_PAGES):
        urls.append({"url": f"{RIGHTMOVE_BASE}&index={i * 24}"})
    return urls


# Regexes used for parsing (module-level to compile once)
# Rightmove uses /properties/ID (e.g. .../properties/51867531#/?channel=COM_LET), NOT /commercial-property-to-let/property-ID
_BASE_URL = "https://www.rightmove.co.uk"
_LINK_RE = re.compile(
    r"(?:https?://(?:www\.)?rightmove\.co\.uk)?/properties/(\d+)[^\s\)\]\"]*",
    re.I,
)
_PRICE_RE = re.compile(r"£[\d,]+(?:\s*pcm)?|POA", re.I)
_SIZE_RE = re.compile(r"[\d,]+(?:\s*[–\-]\s*)?[\d,]*(?:\s*\.\s*\d+)?\s*sq\.\s*ft\.?", re.I)
_PHONE_RE = re.compile(r"Call\s*[\(]?(?:tel:)?([\d\s\-]{10,18})", re.I)
_PROPERTY_PAGE_URL_RE = re.compile(
    r"rightmove\.co\.uk/properties/\d+",
    re.I,
)


def _page_url(item: dict) -> str:
    """Get the crawled page URL from an item."""
    return (
        item.get("url")
        or (item.get("crawl") or {}).get("loadedUrl")
        or ""
    ).strip()


def _extract_one_property_from_text(
    block: str, link: str, default_name: str = "Commercial Property"
) -> dict:
    """Parse price, size, phone, name, address from a text block. link is the property URL."""
    price_m = _PRICE_RE.search(block)
    price = price_m.group(0).strip() if price_m else ""
    size_m = _SIZE_RE.search(block)
    size = size_m.group(0).strip() if size_m else ""
    phone_m = _PHONE_RE.search(block)
    phone = (phone_m.group(1).strip() if phone_m else "").replace(" ", " ").strip()

    lines = [l.strip() for l in block.split("\n") if l.strip()]
    property_name = ""
    address = ""
    for line in lines:
        if not line or line == link or line.startswith("http"):
            continue
        if re.match(r"^£|^POA", line):
            if not price:
                price = line.split()[0] if line.split() else ""
            continue
        if "sq. ft." in line.lower():
            if not size:
                size = line[:80]
            continue
        if "Call" in line or "Local call rate" in line:
            continue
        if "rightmove" in line.lower():
            continue
        if len(line) > 15 and not property_name and len(line) < 250:
            property_name = line[:200]
            continue
        if len(line) > 10 and "," in line and (
            "London" in line or "EC" in line or "W1" in line or "E14" in line or "SW" in line or "NW" in line
        ):
            address = line[:300]
            break

    if not property_name and lines:
        for L in lines:
            if 15 < len(L) < 220 and not L.startswith("http") and "rightmove" not in L.lower():
                property_name = L[:200]
                break

    details = block[:2000].strip() if len(block) > 2000 else block.strip()
    return {
        "Property Name": property_name or default_name,
        "Address": address,
        "Link to Property": link,
        "Price": price,
        "Size": size,
        "Property Details": details,
        "Phone Number": phone,
    }


def parse_properties_from_item(item: dict) -> list[dict]:
    """Extract property rows from one crawler output item (page content).
    Handles: (1) property detail pages (one item = one property), (2) list pages (multiple links in text).
    """
    page_url = _page_url(item)
    text = (
        item.get("markdown")
        or item.get("text")
        or item.get("content")
        or item.get("readableText")
        or ""
    )

    # Case 1: This item is a property detail page (crawler followed a link to /property-12345)
    if _PROPERTY_PAGE_URL_RE.search(page_url) and "/find.html" not in page_url:
        meta = item.get("metadata") or {}
        title = (meta.get("title") or "").strip()
        if title and "rightmove" in title.lower():
            # e.g. "Property Name | Rightmove" -> use first part
            title = title.split("|")[0].strip()
        default_name = title or "Commercial Property"
        row = _extract_one_property_from_text(text, page_url, default_name=default_name)
        row["Property Name"] = row["Property Name"] or default_name
        return [row]

    if not text:
        return []

    # Case 2: List/search page – find all property links and extract a block per link
    rows = []
    links = list(_LINK_RE.finditer(text))
    seen_ids = set()
    for i, m in enumerate(links):
        prop_id = m.group(1)
        if prop_id in seen_ids:
            continue
        seen_ids.add(prop_id)
        raw = m.group(0).strip()
        link = raw if raw.startswith("http") else (_BASE_URL + raw)
        start = m.start()
        end = links[i + 1].start() if i + 1 < len(links) else len(text)
        block = text[start:end]
        rows.append(_extract_one_property_from_text(block, link))

    return rows


def dedupe_rows(rows: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for r in rows:
        key = (r.get("Link to Property") or "").strip()
        if key and key not in seen:
            seen.add(key)
            out.append(r)
    return out


def write_excel(rows: list[dict], path: Path) -> None:
    """Write all rows to Excel (overwrites file)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rightmove London Commercial"
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    for row_idx, r in enumerate(rows, 2):
        for col_idx, h in enumerate(HEADERS, 1):
            val = r.get(h, "")
            if isinstance(val, str) and len(val) > 32767:
                val = val[:32767]
            ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


def main():
    debug = "--debug" in sys.argv
    test_mode = "--test" in sys.argv
    token = (
        os.environ.get("APIFY_TOKEN")
        or os.environ.get("APIFY_API_TOKEN")
        or os.environ.get("Apify_Website_API_KEY")
    )
    if not token:
        log.error("Missing Apify token. Set APIFY_TOKEN, APIFY_API_TOKEN, or Apify_Website_API_KEY in environment or .env")
        sys.exit(1)

    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / "rightmove_london_commercial.xlsx"
    log.info("Output file: %s", out_path)
    client = ApifyClient(token=token)
    start_urls = build_start_urls()

    if test_mode:
        start_urls = start_urls[:2]
        max_pages = 10
        log.info("TEST MODE: 2 start URL(s), max %s pages", max_pages)
        run_input = {
            "startUrls": start_urls,
            "maxCrawlPages": max_pages,
            "maxCrawlDepth": 2,
            "saveMarkdown": True,
            "crawlerType": "playwright:adaptive",
        }
        try:
            log.info("Calling Apify actor (blocking)...")
            run = client.actor("apify/website-content-crawler").call(run_input=run_input)
            dataset_id = run["defaultDatasetId"]
            log.info("Run finished. Dataset ID: %s", dataset_id)
        except Exception as e:
            log.error("Apify run failed: %s", e)
            sys.exit(1)
        log.info("Fetching dataset items...")
        items = list(client.dataset(dataset_id).iterate_items())
        log.info("Received %s page(s) from dataset", len(items))
        if debug:
            def _truncate(o, max_len=2000):
                if isinstance(o, str) and len(o) > max_len:
                    return o[:max_len] + f"... [truncated, total {len(o)} chars]"
                if isinstance(o, dict):
                    return {k: _truncate(v, max_len) for k, v in o.items()}
                if isinstance(o, list):
                    return [_truncate(x, max_len) for x in o[:5]]
                return o
            with open(out_dir / "debug_crawler_items.json", "w", encoding="utf-8") as f:
                json.dump([_truncate(i) for i in items[:10]], f, indent=2, ensure_ascii=False)
            log.info("Debug: saved first 10 items to %s", out_dir / "debug_crawler_items.json")
        all_rows = []
        for item in items:
            all_rows.extend(parse_properties_from_item(item))
        all_rows = dedupe_rows(all_rows)
        write_excel(all_rows, out_path)
        log.info("Extracted %s unique properties. Saved: %s", len(all_rows), out_path)
        return out_path

    # If APIFY_RUN_ID is set: harvest that run's dataset (no new run = no extra memory)
    existing_run_id = os.environ.get("APIFY_RUN_ID", "").strip()
    if existing_run_id:
        log.info("Harvesting existing run: %s (poll every %ss, save every %s properties)", existing_run_id, 45, SAVE_EVERY_N_PROPERTIES)
        run_info = client.run(existing_run_id).get()
        dataset_id = run_info.get("defaultDatasetId")
        if not dataset_id:
            log.error("Run has no defaultDatasetId")
            sys.exit(1)
        log.info("Dataset ID: %s | Run status: %s", dataset_id, run_info.get("status"))
        all_rows = []
        seen_links = set()
        last_saved_count = 0
        last_offset = 0
        poll_interval = 45
        poll_num = 0
        while True:
            time.sleep(poll_interval)
            poll_num += 1
            run_info = client.run(existing_run_id).get()
            status = run_info.get("status")
            page = client.dataset(dataset_id).list_items(offset=last_offset, limit=500)
            items = page.items
            total_in_dataset = getattr(page, "total", last_offset + len(items))
            log.info("Poll #%s | run_status=%s | fetched %s items (offset=%s, dataset_total=%s)", poll_num, status, len(items), last_offset, total_in_dataset)
            for item in items:
                for r in parse_properties_from_item(item):
                    key = (r.get("Link to Property") or "").strip()
                    if key and key not in seen_links:
                        seen_links.add(key)
                        all_rows.append(r)
            last_offset += len(items)
            new_since_save = len(all_rows) - last_saved_count
            if new_since_save >= SAVE_EVERY_N_PROPERTIES or (items and last_saved_count == 0):
                write_excel(all_rows, out_path)
                last_saved_count = len(all_rows)
                log.info("SAVED Excel: %s unique properties -> %s", len(all_rows), out_path)
            if status in ("SUCCEEDED", "FAILED", "ABORTED") and not items:
                log.info("Run finished (status=%s), no more items. Exiting poll loop.", status)
                break
        write_excel(all_rows, out_path)
        log.info("DONE. Total %s unique properties. Final save: %s", len(all_rows), out_path)
        return out_path

    # Full run in chunks (fits Apify memory limit); run each chunk, parse, save every 10, then next chunk
    all_rows = []
    seen_links = set()
    last_saved_count = 0
    num_chunks = (len(start_urls) + URLS_PER_CHUNK - 1) // URLS_PER_CHUNK
    log.info("Full run: %s start URL(s) in %s chunks of %s URLs. Save every %s properties.", len(start_urls), num_chunks, URLS_PER_CHUNK, SAVE_EVERY_N_PROPERTIES)

    for chunk_start in range(0, len(start_urls), URLS_PER_CHUNK):
        chunk_urls = start_urls[chunk_start : chunk_start + URLS_PER_CHUNK]
        chunk_num = chunk_start // URLS_PER_CHUNK + 1
        run_input = {
            "startUrls": chunk_urls,
            "maxCrawlPages": MAX_PAGES_PER_CHUNK,
            "maxCrawlDepth": 2,
            "saveMarkdown": True,
            "crawlerType": "playwright:adaptive",
        }
        run = None
        for attempt in range(1, CHUNK_MAX_RETRIES + 1):
            log.info("--- Chunk %s/%s (%s URLs) --- Attempt %s/%s - Starting Apify actor (blocking)...", chunk_num, num_chunks, len(chunk_urls), attempt, CHUNK_MAX_RETRIES)
            try:
                run = client.actor("apify/website-content-crawler").call(run_input=run_input)
                break
            except Exception as e:
                err_str = str(e)
                if "memory limit" in err_str.lower() or "8192" in err_str:
                    if attempt < CHUNK_MAX_RETRIES:
                        log.warning("Chunk %s failed (Apify memory limit). Waiting %ss then retrying...", chunk_num, CHUNK_RETRY_WAIT_SEC)
                        time.sleep(CHUNK_RETRY_WAIT_SEC)
                    else:
                        log.error("Chunk %s failed after %s retries: %s", chunk_num, CHUNK_MAX_RETRIES, e)
                        write_excel(all_rows, out_path)
                        sys.exit(1)
                else:
                    log.error("Chunk %s failed: %s", chunk_num, e)
                    write_excel(all_rows, out_path)
                    sys.exit(1)
        if run is None:
            log.error("Chunk %s: no run returned", chunk_num)
            write_excel(all_rows, out_path)
            sys.exit(1)
        dataset_id = run["defaultDatasetId"]
        log.info("Chunk %s finished. Fetching dataset items...", chunk_num)
        items = list(client.dataset(dataset_id).iterate_items())
        log.info("Chunk %s: received %s page(s) from dataset", chunk_num, len(items))

        for item in items:
            for r in parse_properties_from_item(item):
                key = (r.get("Link to Property") or "").strip()
                if key and key not in seen_links:
                    seen_links.add(key)
                    all_rows.append(r)

        new_count = len(all_rows) - last_saved_count
        if new_count >= SAVE_EVERY_N_PROPERTIES or len(all_rows) > 0:
            write_excel(all_rows, out_path)
            last_saved_count = len(all_rows)
            log.info("SAVED Excel: %s unique properties -> %s", len(all_rows), out_path)

    write_excel(all_rows, out_path)
    log.info("DONE. Total %s unique properties. Final save: %s", len(all_rows), out_path)
    return out_path


if __name__ == "__main__":
    main()
