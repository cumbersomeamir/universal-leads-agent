"""Export leads to XLSX and JSONL."""

import json
from datetime import datetime, timezone
from pathlib import Path

from core.config import get_config
from core.models import Lead


def _output_dir() -> Path:
    cfg = get_config()
    d = Path(cfg["output_dir"])
    d.mkdir(parents=True, exist_ok=True)
    return d


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def export_xlsx(leads: list[Lead]) -> str:
    """Write leads to XLSX; return path."""
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise RuntimeError("pip install openpyxl")
    cfg = get_config()
    prefix = cfg.get("xlsx_prefix") or "leads_"
    path = _output_dir() / f"{prefix}{_timestamp()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = [
        "client_name", "post_url", "email", "project_description",
        "platform", "post_date", "post_text_snippet", "company", "source_type",
        "confidence_score", "email_source", "keywords_matched", "location",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
    for row, lead in enumerate(leads, 2):
        row_data = lead.to_row()
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=row_data.get(h, ""))
    wb.save(path)
    return str(path)


def export_jsonl(leads: list[Lead]) -> str:
    """Write leads to JSONL; return path."""
    cfg = get_config()
    prefix = cfg.get("jsonl_prefix") or "leads_"
    path = _output_dir() / f"{prefix}{_timestamp()}.jsonl"
    with open(path, "w") as f:
        for lead in leads:
            f.write(json.dumps(lead.model_dump(mode="json")) + "\n")
    return str(path)
